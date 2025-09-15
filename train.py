import os, argparse, time, datetime, stat, shutil
import numpy as np
import torch
import torch.nn.functional as F
from torch.autograd import Variable
from torch.utils.data import DataLoader
import torchvision.utils as vutils
from util.MFE_dataset import MFE_dataset
from util.augmentation import RandomFlip, RandomCrop, RandomCropOut, RandomBrightness, RandomNoise
from util.util import compute_results
from sklearn.metrics import confusion_matrix
from torch.utils.tensorboard import SummaryWriter
from model import model
from openpyxl import Workbook, load_workbook


#############################################################################################
parser = argparse.ArgumentParser(description='Train with pytorch')
parser.add_argument('--model_name', '-m', type=str, default='model')
parser.add_argument('--batch_size', '-b', type=int, default=2)
parser.add_argument('--lr_start', '-ls', type=float, default=0.01)
parser.add_argument('--gpu', '-g', type=int, default=0)
parser.add_argument('--lr_decay', '-ld', type=float, default=0.95)
parser.add_argument('--epoch_max', '-em', type=int, default=200)
parser.add_argument('--epoch_from', '-ef', type=int, default=0)
parser.add_argument('--num_workers', '-j', type=int, default=8)
parser.add_argument('--n_class', '-nc', type=int, default=7)
parser.add_argument('--data_dir', '-dr', type=str, default='E:\hxy\Datasets')
args = parser.parse_args()
#############################################################################################

augmentation_methods = [
    RandomFlip(prob=0.5),
    RandomCrop(crop_rate=0.1, prob=1.0),
]

def auto_weight_bce(y_hat_log, y):
        if y.ndim == 3:
            y = y.unsqueeze(1)
        with torch.no_grad():
            beta = y.mean(dim=[2, 3], keepdims=True)
        logit_1 = F.logsigmoid(y_hat_log)
        logit_0 = F.logsigmoid(-y_hat_log)
        loss = -(1 - beta) * logit_1 * y - beta * logit_0 * (1 - y)

        return loss.mean()

def train(epo, model, train_loader, optimizer):
    model.train()
    total_loss = []#loss
    for it, (images, rgb_labels, edge_labels, names) in enumerate(train_loader):
        images = Variable(images).cuda(args.gpu)
        rgb_labels = Variable(rgb_labels).cuda(args.gpu)
        edge_labels = Variable(edge_labels).cuda(args.gpu)
        start_t = time.time()
        optimizer.zero_grad()
        edge_result,rgb_result = model(images)
        loss1 = F.cross_entropy(rgb_result, rgb_labels)
        total_loss.append(loss1)  # loss
        loss2 = auto_weight_bce(edge_result,edge_labels.float())
        loss = loss1+loss2
        loss.backward()
        optimizer.step()
        lr_this_epo=0
        for param_group in optimizer.param_groups:
            lr_this_epo = param_group['lr']
        print('Train: %s, epo %s/%s, iter %s/%s, lr %.8f, %.2f img/sec, loss %.4f, loss1 %.4f, loss2 %.4f  time %s' \
            % (args.model_name, epo, args.epoch_max, it+1, len(train_loader), lr_this_epo, len(names)/(time.time()-start_t), float(loss), float(loss1),float(loss2),
              datetime.datetime.now().replace(microsecond=0)-start_datetime))
        if accIter['train'] % 1 == 0:
            writer.add_scalar('Train/loss', loss, accIter['train'])
            writer.add_scalar('Train/loss1', loss1, accIter['train'])
            writer.add_scalar('Train/loss2', loss2, accIter['train'])
        view_figure = True
        if accIter['train'] % 20 == 0:
            if view_figure:
                input_rgb_images = vutils.make_grid(images[:,:3], nrow=8, padding=10)
                writer.add_image('Train/input_rgb_images', input_rgb_images, accIter['train'])
                scale = max(1, 255//args.n_class)
                groundtruth_tensor = rgb_labels.unsqueeze(1) * scale
                groundtruth_tensor = torch.cat((groundtruth_tensor, groundtruth_tensor, groundtruth_tensor), 1)
                groudtruth_images = vutils.make_grid(groundtruth_tensor, nrow=8, padding=10)
                writer.add_image('Train/groudtruth_images', groudtruth_images, accIter['train'])
                predicted_tensor = rgb_result.argmax(1).unsqueeze(1) * scale
                predicted_tensor = torch.cat((predicted_tensor, predicted_tensor, predicted_tensor),1)
                predicted_images = vutils.make_grid(predicted_tensor, nrow=8, padding=10)
                writer.add_image('Train/predicted_images', predicted_images, accIter['train'])
                edgetruth_tensor = edge_labels.unsqueeze(1) * scale
                edgetruth_tensor = torch.cat((edgetruth_tensor, edgetruth_tensor, edgetruth_tensor), 1)
                edgetruth_images = vutils.make_grid(edgetruth_tensor, nrow=8, padding=10)
                writer.add_image('Train/edgetruth_images', edgetruth_images, accIter['train'])
                edge_predicted_tensor = edge_result
                edge_predicted_tensor = torch.cat((edge_predicted_tensor, edge_predicted_tensor, edge_predicted_tensor),1)
                edge_predicted_images = vutils.make_grid(edge_predicted_tensor, nrow=8, padding=10)
                writer.add_image('Train/edge_predicted_images', edge_predicted_images, accIter['train'])

        accIter['train'] = accIter['train'] + 1
    excel_path = "total_loss.xlsx"
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["epoch", "iter_loss"])
    for row in total_loss:
        ws.append([epo, row.item()])
    wb.save(excel_path)

def validation(epo, model, val_loader): 
    model.eval()
    with torch.no_grad():
        total_loss1 = []#loss
        for it, (images, rgb_labels, edge_labels, names) in enumerate(val_loader):
            images = Variable(images).cuda(args.gpu)
            rgb_labels = Variable(rgb_labels).cuda(args.gpu)
            edge_labels = Variable(edge_labels).cuda(args.gpu)
            start_t = time.time()
            edge_result,rgb_result = model(images)
            loss2 = auto_weight_bce(edge_result,edge_labels.float())
            total_loss1.append(loss2)#loss

            print('Val: %s, epo %s/%s, iter %s/%s, %.2f img/sec, loss %.4f, time %s' \
                  % (args.model_name, epo, args.epoch_max, it + 1, len(val_loader), len(names)/(time.time()-start_t), float(loss2),
                    datetime.datetime.now().replace(microsecond=0)-start_datetime))
            if accIter['val'] % 1 == 0:
                writer.add_scalar('Validation/loss', loss2, accIter['val'])
            view_figure = False
            if accIter['val'] % 100 == 0:
                if view_figure:
                    input_rgb_images = vutils.make_grid(images[:, :3], nrow=8, padding=10)
                    writer.add_image('Validation/input_rgb_images', input_rgb_images, accIter['val'])
                    scale = max(1, 255 // args.n_class)
                    groundtruth_tensor = rgb_labels.unsqueeze(1) * scale
                    groundtruth_tensor = torch.cat((groundtruth_tensor, groundtruth_tensor, groundtruth_tensor), 1)
                    groudtruth_images = vutils.make_grid(groundtruth_tensor, nrow=8, padding=10)
                    writer.add_image('Validation/groudtruth_images', groudtruth_images, accIter['val'])
                    predicted_tensor = rgb_result.argmax(1).unsqueeze(1)*scale
                    predicted_tensor = torch.cat((predicted_tensor, predicted_tensor, predicted_tensor), 1)
                    predicted_images = vutils.make_grid(predicted_tensor, nrow=8, padding=10)
                    writer.add_image('Validation/predicted_images', predicted_images, accIter['val'])
            accIter['val'] += 1
        excel_path = "total_loss1.xlsx"
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["epoch", "val_loss"])
        for row in total_loss1:
            ws.append([epo, row.item()])
        wb.save(excel_path)

def testing(epo, model, test_loader):
    model.eval()
    conf_total = np.zeros((args.n_class, args.n_class))
    label_list = ["unlabeled", "reban", "yinying", "yiwu", "niaofen", "duanlu", "tree"]
    testing_results_file = os.path.join(weight_dir, 'testing_results_file.txt')
    with torch.no_grad():
        for it, (images, rgb_labels, edge_labels,names) in enumerate(test_loader):
            images = Variable(images).cuda(args.gpu)
            rgb_labels = Variable(rgb_labels).cuda(args.gpu)
            edge_labels = Variable(edge_labels).cuda(args.gpu)
            edge_result,rgb_result = model(images)
            rgb_labels = rgb_labels.cpu().numpy().squeeze().flatten()
            prediction = rgb_result.argmax(1).cpu().numpy().squeeze().flatten()
            conf = confusion_matrix(y_true=rgb_labels, y_pred=prediction, labels=[0,1,2,3,4,5,6])
            conf_total += conf
            print('Test: %s, epo %s/%s, iter %s/%s, time %s' % (args.model_name, epo, args.epoch_max, it+1, len(test_loader),
                 datetime.datetime.now().replace(microsecond=0)-start_datetime))
    precision, recall, IoU = compute_results(conf_total)
    writer.add_scalar('Test/average_recall', recall.mean(), epo)
    writer.add_scalar('Test/average_IoU', IoU.mean(), epo)
    writer.add_scalar('Test/average_precision',precision.mean(), epo)
    for i in range(len(precision)):
        writer.add_scalar("Test(class)/precision_class_%s" % label_list[i], precision[i], epo)
        writer.add_scalar("Test(class)/recall_class_%s"% label_list[i], recall[i],epo)
        writer.add_scalar('Test(class)/Iou_%s'% label_list[i], IoU[i], epo)
    if epo==0:
        with open(testing_results_file, 'w') as f:
            f.write("# %s, initial lr: %s, batch size: %s, date: %s \n" %(args.model_name, args.lr_start, args.batch_size, datetime.date.today()))
            f.write("# epoch: unlabeled, reban, yinying, yiwu, niaofen, duanlu, tree, average(nan_to_num). (Acc %, IoU %)\n")
    with open(testing_results_file, 'a') as f:
        f.write(str(epo)+': ')
        for i in range(len(precision)):
            f.write('%0.4f, %0.4f ' % (100*recall[i], 100*IoU[i]))
        f.write('%0.4f, %0.4f\n' % (100*np.mean(np.nan_to_num(recall)), 100*np.mean(np.nan_to_num(IoU))))
    print('saving testing results.')
    with open(testing_results_file, "r") as file:
        writer.add_text('testing_results', file.read().replace('\n', '  \n'), epo)

if __name__ == '__main__':
   
    torch.cuda.set_device(args.gpu)
    print("\nthe pytorch version:", torch.__version__)
    print("the gpu count:", torch.cuda.device_count())
    print("the current used gpu:", torch.cuda.current_device(), '\n')

    model = eval(args.model_name)(n_class=args.n_class, training_style='train')

    if args.gpu >= 0: model.cuda(args.gpu)
    optimizer = torch.optim.SGD(model.parameters(), lr=args.lr_start, momentum=0.9, weight_decay=0.0005)
    scheduler = torch.optim.lr_scheduler.ExponentialLR(optimizer, gamma=args.lr_decay, last_epoch=-1)

    weight_dir = os.path.join("./runst", args.model_name)

    writer = SummaryWriter("./runst/tensorboard_log")

    print('training %s on GPU #%d with pytorch' % (args.model_name, args.gpu))
    print('from epoch %d / %s' % (args.epoch_from, args.epoch_max))
    print('weight will be saved in: %s' % weight_dir)

    train_dataset = MFE_dataset(data_dir=args.data_dir, split='train')
    val_dataset  = MFE_dataset(data_dir=args.data_dir, split='val')
    test_dataset = MFE_dataset(data_dir=args.data_dir, split='test')

    train_loader  = DataLoader(
        dataset     = train_dataset,
        batch_size  = args.batch_size,
        shuffle     = True,
        num_workers = args.num_workers,
        pin_memory  = True,
        drop_last   = False
    )
    val_loader  = DataLoader(
        dataset     = val_dataset,
        batch_size  = args.batch_size,
        shuffle     = False,
        num_workers = args.num_workers,
        pin_memory  = True,
        drop_last   = False
    )
    test_loader = DataLoader(
        dataset      = test_dataset,
        batch_size   = args.batch_size,
        shuffle      = False,
        num_workers  = args.num_workers,
        pin_memory   = True,
        drop_last    = False
    )
    start_datetime = datetime.datetime.now().replace(microsecond=0)
    accIter = {'train': 0, 'val': 0}
    for epo in range(args.epoch_from, args.epoch_max):
        print('\ntrain %s, epo #%s begin...' % (args.model_name, epo))

        train(epo, model, train_loader, optimizer)
        validation(epo, model, val_loader)

        checkpoint_model_file = os.path.join(weight_dir, str(epo) + '.pth')
        print('saving check point %s: ' % checkpoint_model_file)
        torch.save(model.state_dict(), checkpoint_model_file)

        testing(epo, model, test_loader)
        scheduler.step()
